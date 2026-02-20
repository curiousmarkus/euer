import csv
import sys
from pathlib import Path

from ..config import get_export_dir, load_config
from ..constants import DEFAULT_EXPORT_DIR
from ..db import get_db_connection

# Optional: openpyxl für XLSX-Export
try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def cmd_export(args):
    """Exportiert Daten als CSV oder XLSX."""
    db_path = Path(args.db)
    conn = get_db_connection(db_path)

    config = load_config()
    config_export_dir = get_export_dir(config)

    if args.output:
        output_dir = Path(args.output)
    elif config_export_dir:
        output_dir = Path(config_export_dir)
    else:
        output_dir = DEFAULT_EXPORT_DIR

    output_dir.mkdir(exist_ok=True)

    year = args.year

    if year is not None:
        year_filter = "WHERE strftime('%Y', e.date) = ?"
        year_params = (str(year),)
        income_filter = "WHERE strftime('%Y', i.date) = ?"
        income_params = (str(year),)
        private_filter = "WHERE strftime('%Y', p.date) = ?"
        private_params = (str(year),)
        sacheinlagen_filter = "WHERE e.is_private_paid = 1 AND strftime('%Y', e.date) = ?"
        sacheinlagen_params = (str(year),)
    else:
        year_filter = ""
        year_params = ()
        income_filter = ""
        income_params = ()
        private_filter = ""
        private_params = ()
        sacheinlagen_filter = "WHERE e.is_private_paid = 1"
        sacheinlagen_params = ()

    has_private_transfers = (
        conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='private_transfers'"
        ).fetchone()
        is not None
    )
    expense_columns = {
        row["name"]
        for row in conn.execute("PRAGMA table_info(expenses)").fetchall()
    }
    has_private_expense_cols = {
        "is_private_paid",
        "private_classification",
    }.issubset(expense_columns)

    # Ausgaben laden
    expenses = conn.execute(
        f"""SELECT e.receipt_name, e.date, e.vendor, c.name as category, c.eur_line,
                  e.amount_eur, e.account, e.foreign_amount, e.notes, e.is_rc, e.vat_input, e.vat_output
           FROM expenses e
           LEFT JOIN categories c ON e.category_id = c.id
           {year_filter}
           ORDER BY e.date, e.id""",
        year_params,
    ).fetchall()

    # Einnahmen laden
    income = conn.execute(
        f"""SELECT i.receipt_name, i.date, i.source, c.name as category, c.eur_line,
                  i.amount_eur, i.foreign_amount, i.notes, i.vat_output
           FROM income i
           LEFT JOIN categories c ON i.category_id = c.id
           {income_filter}
           ORDER BY i.date, i.id""",
        income_params,
    ).fetchall()

    if has_private_transfers:
        private_transfers = conn.execute(
            f"""SELECT p.id, p.date, p.type, p.amount_eur, p.description,
                      p.notes, p.related_expense_id
               FROM private_transfers p
               {private_filter}
               ORDER BY p.date, p.id""",
            private_params,
        ).fetchall()
    else:
        private_transfers = []

    if has_private_expense_cols:
        sacheinlagen = conn.execute(
            f"""SELECT e.id, e.date, e.vendor, c.name as category, e.amount_eur,
                      e.account, e.private_classification
               FROM expenses e
               LEFT JOIN categories c ON e.category_id = c.id
               {sacheinlagen_filter}
               ORDER BY e.date, e.id""",
            sacheinlagen_params,
        ).fetchall()
    else:
        sacheinlagen = []

    conn.close()

    if args.format == "csv":
        # CSV Export
        exp_suffix = f"_{year}" if year is not None else ""
        exp_path = output_dir / f"EÜR{exp_suffix}_Ausgaben.csv"
        with open(exp_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "Belegname",
                    "Datum",
                    "Lieferant",
                    "Kategorie",
                    "EUR",
                    "Konto",
                    "Fremdwährung",
                    "Bemerkung",
                    "RC",
                    "Vorsteuer",
                    "Umsatzsteuer",
                ]
            )
            for r in expenses:
                if r["category"]:
                    cat = (
                        f"{r['category']} ({r['eur_line']})"
                        if r["eur_line"]
                        else r["category"]
                    )
                else:
                    cat = "Ohne Kategorie"
                writer.writerow(
                    [
                        r["receipt_name"] or "",
                        r["date"],
                        r["vendor"],
                        cat,
                        f"{r['amount_eur']:.2f}",
                        r["account"] or "",
                        r["foreign_amount"] or "",
                        r["notes"] or "",
                        "X" if r["is_rc"] else "",
                        f"{r['vat_input']:.2f}" if r["vat_input"] else "",
                        f"{r['vat_output']:.2f}" if r["vat_output"] else "",
                    ]
                )
        print(f"Exportiert: {exp_path}")

        inc_path = output_dir / f"EÜR{exp_suffix}_Einnahmen.csv"
        with open(inc_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "Belegname",
                    "Datum",
                    "Quelle",
                    "Kategorie",
                    "EUR",
                    "Fremdwährung",
                    "Bemerkung",
                    "Umsatzsteuer",
                ]
            )
            for r in income:
                if r["category"]:
                    cat = (
                        f"{r['category']} ({r['eur_line']})"
                        if r["eur_line"]
                        else r["category"]
                    )
                else:
                    cat = "Ohne Kategorie"
                writer.writerow(
                    [
                        r["receipt_name"] or "",
                        r["date"],
                        r["source"],
                        cat,
                        f"{r['amount_eur']:.2f}",
                        r["foreign_amount"] or "",
                        r["notes"] or "",
                        f"{r['vat_output']:.2f}" if r["vat_output"] else "",
                    ]
                )
        print(f"Exportiert: {inc_path}")

        private_path = output_dir / f"EÜR{exp_suffix}_PrivateTransfers.csv"
        with open(private_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "ID",
                    "Datum",
                    "Typ",
                    "EUR",
                    "Beschreibung",
                    "Bemerkung",
                    "related_expense_id",
                ]
            )
            for r in private_transfers:
                writer.writerow(
                    [
                        r["id"],
                        r["date"],
                        r["type"],
                        f"{r['amount_eur']:.2f}",
                        r["description"],
                        r["notes"] or "",
                        r["related_expense_id"] or "",
                    ]
                )
        print(f"Exportiert: {private_path}")

        sache_path = output_dir / f"EÜR{exp_suffix}_Sacheinlagen.csv"
        with open(sache_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "expense_id",
                    "Datum",
                    "Lieferant",
                    "Kategorie",
                    "EUR",
                    "Konto",
                    "Klassifikation",
                ]
            )
            for r in sacheinlagen:
                writer.writerow(
                    [
                        r["id"],
                        r["date"],
                        r["vendor"],
                        r["category"] or "",
                        f"{abs(r['amount_eur']):.2f}",
                        r["account"] or "",
                        r["private_classification"] or "",
                    ]
                )
        print(f"Exportiert: {sache_path}")

    else:
        # XLSX Export
        if not HAS_OPENPYXL:
            print(
                "Fehler: openpyxl nicht installiert. Bitte 'pip install openpyxl'.",
                file=sys.stderr,
            )
            sys.exit(1)

        # Ausgaben
        exp_suffix = f"_{year}" if year is not None else ""
        exp_path = output_dir / f"EÜR{exp_suffix}_Ausgaben.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ausgaben"
        ws.append(
            [
                "Belegname",
                "Datum",
                "Lieferant",
                "Kategorie",
                "EUR",
                "Konto",
                "Fremdwährung",
                "Bemerkung",
                "RC",
                "USt-VA",
            ]
        )
        for r in expenses:
            if r["category"]:
                cat = (
                    f"{r['category']} ({r['eur_line']})"
                    if r["eur_line"]
                    else r["category"]
                )
            else:
                cat = "Ohne Kategorie"
            ws.append(
                [
                    r["receipt_name"] or "",
                    r["date"],
                    r["vendor"],
                    cat,
                    r["amount_eur"],
                    r["account"] or "",
                    r["foreign_amount"] or "",
                    r["notes"] or "",
                    "X" if r["is_rc"] else "",
                    r["vat_output"] if r["vat_output"] else None,
                ]
            )
        wb.save(exp_path)
        print(f"Exportiert: {exp_path}")

        # Einnahmen
        inc_path = output_dir / f"EÜR{exp_suffix}_Einnahmen.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Einnahmen"
        ws.append(
            [
                "Belegname",
                "Datum",
                "Quelle",
                "Kategorie",
                "EUR",
                "Fremdwährung",
                "Bemerkung",
            ]
        )
        for r in income:
            if r["category"]:
                cat = (
                    f"{r['category']} ({r['eur_line']})"
                    if r["eur_line"]
                    else r["category"]
                )
            else:
                cat = "Ohne Kategorie"
            ws.append(
                [
                    r["receipt_name"] or "",
                    r["date"],
                    r["source"],
                    cat,
                    r["amount_eur"],
                    r["foreign_amount"] or "",
                    r["notes"] or "",
                ]
            )
        wb.save(inc_path)
        print(f"Exportiert: {inc_path}")

        private_path = output_dir / f"EÜR{exp_suffix}_Privatvorgaenge.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PrivateTransfers"
        ws.append(
            [
                "ID",
                "Datum",
                "Typ",
                "EUR",
                "Beschreibung",
                "Bemerkung",
                "related_expense_id",
            ]
        )
        for r in private_transfers:
            ws.append(
                [
                    r["id"],
                    r["date"],
                    r["type"],
                    r["amount_eur"],
                    r["description"],
                    r["notes"] or "",
                    r["related_expense_id"] or None,
                ]
            )

        ws2 = wb.create_sheet("Sacheinlagen")
        ws2.append(
            [
                "expense_id",
                "Datum",
                "Lieferant",
                "Kategorie",
                "EUR",
                "Konto",
                "Klassifikation",
            ]
        )
        for r in sacheinlagen:
            ws2.append(
                [
                    r["id"],
                    r["date"],
                    r["vendor"],
                    r["category"] or "",
                    abs(r["amount_eur"]),
                    r["account"] or "",
                    r["private_classification"] or "",
                ]
            )
        wb.save(private_path)
        print(f"Exportiert: {private_path}")
