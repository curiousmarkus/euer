#!/usr/bin/env python3
"""
Excel-Migration für EÜR-Ledger

Importiert bestehende Excel-Daten (2025/2026) in die SQLite-Datenbank.
"""

import argparse
import hashlib
import json
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(
        "Fehler: openpyxl nicht installiert. Bitte 'pip install openpyxl'.",
        file=sys.stderr,
    )
    sys.exit(1)

# =============================================================================
# Konfiguration
# =============================================================================

DEFAULT_DB_PATH = Path(__file__).parent / "euer.db"
DEFAULT_USER = "markus"

# Spalten-Mapping für Excel-Sheets
# Ausgaben: A=Belegname, B=Datum, C=Lieferant, D=Kategorie, E=EUR, F=Konto, G=Fremdwährung, H=Bemerkung, I=USt-VA
# Einnahmen: A=Belegname, B=Datum, C=Quelle, D=Kategorie, E=EUR, F=Fremdwährung, G=Bemerkung

# =============================================================================
# Hilfsfunktionen
# =============================================================================


def compute_hash(
    date: str, vendor_or_source: str, amount_eur: float, receipt_name: str = ""
) -> str:
    """Erzeugt einen eindeutigen Hash für eine Transaktion."""
    data = f"{date}|{vendor_or_source}|{amount_eur:.2f}|{receipt_name or ''}"
    return hashlib.sha256(data.encode("utf-8")).hexdigest()


def get_db_connection(db_path: Path) -> sqlite3.Connection:
    """Erstellt eine Datenbankverbindung mit Row-Factory."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def log_audit(
    conn: sqlite3.Connection,
    table_name: str,
    record_id: int,
    action: str,
    old_data: dict | None = None,
    new_data: dict | None = None,
    user: str = DEFAULT_USER,
    source_file: str | None = None,
):
    """Schreibt einen Audit-Log-Eintrag."""
    if new_data and source_file:
        new_data = {**new_data, "_source": source_file}

    conn.execute(
        """INSERT INTO audit_log (table_name, record_id, action, old_data, new_data, user)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (
            table_name,
            record_id,
            action,
            json.dumps(old_data, ensure_ascii=False) if old_data else None,
            json.dumps(new_data, ensure_ascii=False) if new_data else None,
            user,
        ),
    )


def parse_category(category_str: str) -> tuple[str, int | None]:
    """
    Extrahiert Kategorie-Name und EÜR-Zeile aus String wie "Laufende EDV-Kosten (51)".

    Returns: (name, eur_line) - eur_line kann None sein
    """
    if not category_str:
        return ("", None)

    # Pattern: "Name (Zahl)" oder "Name"
    match = re.match(r"^(.+?)\s*\((\d+)\)\s*$", category_str.strip())
    if match:
        return (match.group(1).strip(), int(match.group(2)))
    return (category_str.strip(), None)


def get_category_id(conn: sqlite3.Connection, name: str, cat_type: str) -> int | None:
    """Sucht Kategorie-ID nach Name (case-insensitive)."""
    row = conn.execute(
        "SELECT id FROM categories WHERE LOWER(name) = LOWER(?) AND type = ?",
        (name, cat_type),
    ).fetchone()
    return row["id"] if row else None


def format_date(value) -> str | None:
    """Konvertiert Excel-Datum zu ISO-Format."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        # Versuche verschiedene Formate
        for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"]:
            try:
                return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return str(value)


def parse_amount(value) -> float | None:
    """Konvertiert Excel-Betrag zu Float."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Deutsches Format: 1.234,56 -> 1234.56
        cleaned = value.replace(".", "").replace(",", ".").strip()
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def cell_value(row, idx: int) -> str | None:
    """Holt Zellwert oder None."""
    if idx >= len(row):
        return None
    val = row[idx].value
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if isinstance(val, str):
        return val.strip()
    return val


# =============================================================================
# Migration
# =============================================================================


def migrate_expenses(
    conn: sqlite3.Connection, ws, source_file: str, dry_run: bool = False
) -> tuple[int, int]:
    """
    Migriert Ausgaben-Sheet.

    Returns: (imported, skipped)
    """
    imported = 0
    skipped = 0
    errors = []

    rows = list(ws.iter_rows(min_row=2))  # Skip header

    for row_num, row in enumerate(rows, start=2):
        # Leere Zeilen überspringen
        if not any(cell.value for cell in row[:5]):
            continue

        receipt_name = cell_value(row, 0)  # A
        date_val = cell_value(row, 1)  # B
        vendor = cell_value(row, 2)  # C
        category_str = cell_value(row, 3)  # D
        amount_val = cell_value(row, 4)  # E
        account = cell_value(row, 5)  # F
        foreign_amount = cell_value(row, 6)  # G
        notes = cell_value(row, 7)  # H
        vat_val = cell_value(row, 8)  # I

        # Validierung
        date = format_date(date_val)
        if not date:
            errors.append(f"Zeile {row_num}: Ungültiges Datum '{date_val}'")
            continue

        if not vendor:
            errors.append(f"Zeile {row_num}: Kein Lieferant angegeben")
            continue

        amount = parse_amount(amount_val)
        if amount is None:
            errors.append(f"Zeile {row_num}: Ungültiger Betrag '{amount_val}'")
            continue

        vat_amount = parse_amount(vat_val)

        # Kategorie finden
        cat_name, _ = parse_category(str(category_str) if category_str else "")
        cat_id = get_category_id(conn, cat_name, "expense")

        if not cat_id:
            errors.append(f"Zeile {row_num}: Kategorie '{cat_name}' nicht gefunden")
            # Fallback auf "Übrige Betriebsausgaben"
            cat_id = get_category_id(conn, "Übrige Betriebsausgaben", "expense")
            if not cat_id:
                continue

        # Hash berechnen
        tx_hash = compute_hash(date, vendor, amount, receipt_name or "")

        # Duplikat-Prüfung
        existing = conn.execute(
            "SELECT id FROM expenses WHERE hash = ?", (tx_hash,)
        ).fetchone()
        if existing:
            skipped += 1
            continue

        if dry_run:
            print(f"  [DRY] {date} | {vendor[:30]:<30} | {amount:>10.2f} | {cat_name}")
            imported += 1
            continue

        # Insert
        cursor = conn.execute(
            """INSERT INTO expenses 
               (receipt_name, date, vendor, category_id, amount_eur, account, 
                foreign_amount, notes, vat_amount, hash)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                receipt_name,
                date,
                vendor,
                cat_id,
                amount,
                account,
                foreign_amount,
                notes,
                vat_amount,
                tx_hash,
            ),
        )
        record_id = cursor.lastrowid
        assert record_id is not None

        # Audit-Log
        new_data = {
            "receipt_name": receipt_name,
            "date": date,
            "vendor": vendor,
            "category_id": cat_id,
            "amount_eur": amount,
            "account": account,
            "foreign_amount": foreign_amount,
            "notes": notes,
            "vat_amount": vat_amount,
        }
        log_audit(
            conn,
            "expenses",
            record_id,
            "MIGRATE",
            new_data=new_data,
            source_file=source_file,
        )

        imported += 1

    # Fehler ausgeben
    for err in errors:
        print(f"  WARNUNG: {err}", file=sys.stderr)

    return imported, skipped


def migrate_income(
    conn: sqlite3.Connection, ws, source_file: str, dry_run: bool = False
) -> tuple[int, int]:
    """
    Migriert Einnahmen-Sheet.

    Returns: (imported, skipped)
    """
    imported = 0
    skipped = 0
    errors = []

    rows = list(ws.iter_rows(min_row=2))  # Skip header

    for row_num, row in enumerate(rows, start=2):
        # Leere Zeilen überspringen
        if not any(cell.value for cell in row[:5]):
            continue

        receipt_name = cell_value(row, 0)  # A
        date_val = cell_value(row, 1)  # B
        source = cell_value(row, 2)  # C
        category_str = cell_value(row, 3)  # D
        amount_val = cell_value(row, 4)  # E
        foreign_amount = cell_value(row, 5)  # F
        notes = cell_value(row, 6)  # G

        # Validierung
        date = format_date(date_val)
        if not date:
            errors.append(f"Zeile {row_num}: Ungültiges Datum '{date_val}'")
            continue

        if not source:
            errors.append(f"Zeile {row_num}: Keine Quelle angegeben")
            continue

        amount = parse_amount(amount_val)
        if amount is None:
            errors.append(f"Zeile {row_num}: Ungültiger Betrag '{amount_val}'")
            continue

        # Kategorie finden
        cat_name, _ = parse_category(str(category_str) if category_str else "")
        cat_id = get_category_id(conn, cat_name, "income")

        if not cat_id:
            errors.append(f"Zeile {row_num}: Kategorie '{cat_name}' nicht gefunden")
            # Fallback
            cat_id = get_category_id(conn, "Sonstige betriebsfremde Einnahme", "income")
            if not cat_id:
                continue

        # Hash berechnen
        tx_hash = compute_hash(date, source, amount, receipt_name or "")

        # Duplikat-Prüfung
        existing = conn.execute(
            "SELECT id FROM income WHERE hash = ?", (tx_hash,)
        ).fetchone()
        if existing:
            skipped += 1
            continue

        if dry_run:
            print(f"  [DRY] {date} | {source[:30]:<30} | {amount:>10.2f} | {cat_name}")
            imported += 1
            continue

        # Insert
        cursor = conn.execute(
            """INSERT INTO income 
               (receipt_name, date, source, category_id, amount_eur, foreign_amount, notes, hash)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                receipt_name,
                date,
                source,
                cat_id,
                amount,
                foreign_amount,
                notes,
                tx_hash,
            ),
        )
        record_id = cursor.lastrowid
        assert record_id is not None

        # Audit-Log
        new_data = {
            "receipt_name": receipt_name,
            "date": date,
            "source": source,
            "category_id": cat_id,
            "amount_eur": amount,
            "foreign_amount": foreign_amount,
            "notes": notes,
        }
        log_audit(
            conn,
            "income",
            record_id,
            "MIGRATE",
            new_data=new_data,
            source_file=source_file,
        )

        imported += 1

    # Fehler ausgeben
    for err in errors:
        print(f"  WARNUNG: {err}", file=sys.stderr)

    return imported, skipped


def migrate_excel(excel_path: Path, db_path: Path, dry_run: bool = False):
    """Migriert eine Excel-Datei."""
    print(f"Öffne: {excel_path}")

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"Fehler beim Öffnen: {e}", file=sys.stderr)
        sys.exit(1)

    source_file = excel_path.name

    conn = get_db_connection(db_path)

    # Ausgaben
    if "Ausgaben" in wb.sheetnames:
        print("\nMigriere Ausgaben...")
        ws = wb["Ausgaben"]
        exp_imported, exp_skipped = migrate_expenses(conn, ws, source_file, dry_run)
        print(f"  -> {exp_imported} importiert, {exp_skipped} übersprungen (Duplikate)")
    else:
        print("WARNUNG: Sheet 'Ausgaben' nicht gefunden", file=sys.stderr)
        exp_imported, exp_skipped = 0, 0

    # Einnahmen
    if "Einnahmen" in wb.sheetnames:
        print("\nMigriere Einnahmen...")
        ws = wb["Einnahmen"]
        inc_imported, inc_skipped = migrate_income(conn, ws, source_file, dry_run)
        print(f"  -> {inc_imported} importiert, {inc_skipped} übersprungen (Duplikate)")
    else:
        print("WARNUNG: Sheet 'Einnahmen' nicht gefunden", file=sys.stderr)
        inc_imported, inc_skipped = 0, 0

    if not dry_run:
        conn.commit()

    conn.close()
    wb.close()

    print(f"\n{'[DRY RUN] ' if dry_run else ''}Migration abgeschlossen:")
    print(f"  Ausgaben:  {exp_imported} neu, {exp_skipped} Duplikate")
    print(f"  Einnahmen: {inc_imported} neu, {inc_skipped} Duplikate")


# =============================================================================
# Hauptprogramm
# =============================================================================


def main():
    parser = argparse.ArgumentParser(
        description="Migriert Excel-EÜR-Dateien in die SQLite-Datenbank"
    )
    parser.add_argument("excel_file", type=Path, help="Pfad zur Excel-Datei")
    parser.add_argument(
        "--db",
        type=Path,
        default=DEFAULT_DB_PATH,
        help=f"Pfad zur Datenbank (default: {DEFAULT_DB_PATH})",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Zeigt was importiert würde, ohne zu schreiben",
    )

    args = parser.parse_args()

    if not args.excel_file.exists():
        print(f"Fehler: Datei nicht gefunden: {args.excel_file}", file=sys.stderr)
        sys.exit(1)

    if not args.db.exists():
        print(f"Fehler: Datenbank nicht gefunden: {args.db}", file=sys.stderr)
        print("Bitte zuerst 'python euer.py init' ausführen.", file=sys.stderr)
        sys.exit(1)

    migrate_excel(args.excel_file, args.db, args.dry_run)


if __name__ == "__main__":
    main()
